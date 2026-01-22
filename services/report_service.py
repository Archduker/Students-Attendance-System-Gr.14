"""
Report Service - Export Reports
================================

Service xử lý export báo cáo sang các format:
- PDF
- Excel (XLSX)
- CSV
"""

from typing import Dict, Any, List
import csv
from datetime import datetime
from pathlib import Path


class ReportService:
    """
    Service export báo cáo.
    
    Features:
        - Export to PDF
        - Export to Excel
        - Export to CSV
        - Format và style documents
        
    Example:
        >>> report_service = ReportService()
        >>> report_service.export_to_csv(data, "report.csv")
        True
    """
    
    def __init__(self):
        """Khởi tạo ReportService."""
        pass
    
    def export_to_csv(self, data: Dict[str, Any], filename: str) -> bool:
        """
        Export report to CSV.
        
        Args:
            data: Report data dict
            filename: Output filename
            
        Returns:
            True if successful
        """
        try:
            with open(filename, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                
                # Header
                writer.writerow(['Attendance Report'])
                writer.writerow(['Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
                writer.writerow([])
                
                # Summary
                writer.writerow(['Summary'])
                summary = data.get('summary', {})
                for key, value in summary.items():
                    writer.writerow([key, value])
                
                writer.writerow([])
                
                # Details
                writer.writerow(['Details'])
                details = data.get('details', [])
                for item in details:
                    writer.writerow([item])
            
            return True
        except Exception as e:
            print(f"Error exporting to CSV: {e}")
            return False
    
    def export_to_excel(self, data: Dict[str, Any], filename: str) -> bool:
        """
        Export report to Excel (XLSX).
        
        Args:
            data: Report data dict
            filename: Output filename
            
        Returns:
            True if successful
        """
        try:
            # Try to import openpyxl
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, PatternFill
            except ImportError:
                print("Warning: openpyxl not installed. Cannot export to Excel.")
                print("Please install: pip install openpyxl")
                return False
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Attendance Report"
            
            # Title
            ws['A1'] = 'Attendance Report'
            ws['A1'].font = Font(size=16, bold=True)
            
            ws['A2'] = 'Generated:'
            ws['B2'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # Summary section
            row = 4
            ws[f'A{row}'] = 'Summary'
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            summary = data.get('summary', {})
            for key, value in summary.items():
                ws[f'A{row}'] = key
                ws[f'B{row}'] = str(value)
                row += 1
            
            row += 1
            
            # Details section
            ws[f'A{row}'] = 'Details'
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            details = data.get('details', [])
            for item in details:
                ws[f'A{row}'] = str(item)
                row += 1
            
            # Save workbook
            wb.save(filename)
            return True
            
        except Exception as e:
            print(f"Error exporting to Excel: {e}")
            return False
    
    def export_to_pdf(self, data: Dict[str, Any], filename: str) -> bool:
        """
        Export report to PDF.
        
        Args:
            data: Report data dict
            filename: Output filename
            
        Returns:
            True if successful
        """
        try:
            # Try to import reportlab
            try:
                from reportlab.lib.pagesizes import letter, A4
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib.units import inch
                from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
                from reportlab.lib import colors
            except ImportError:
                print("Warning: reportlab not installed. Cannot export to PDF.")
                print("Please install: pip install reportlab")
                return False
            
            # Create PDF
            doc = SimpleDocTemplate(filename, pagesize=letter)
            story = []
            styles = getSampleStyleSheet()
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#1f6aa5'),
                spaceAfter=30,
                alignment=1  # Center
            )
            story.append(Paragraph("Attendance Report", title_style))
            
            # Generated date
            date_text = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            story.append(Paragraph(date_text, styles['Normal']))
            story.append(Spacer(1, 0.3*inch))
            
            # Summary section
            story.append(Paragraph("Summary", styles['Heading2']))
            summary = data.get('summary', {})
            
            summary_data = [[key, str(value)] for key, value in summary.items()]
            if summary_data:
                summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
                summary_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f0f0f0')),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 11),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                    ('GRID', (0, 0), (-1, -1), 1, colors.grey)
                ]))
                story.append(summary_table)
            
            story.append(Spacer(1, 0.3*inch))
            
            # Details section
            story.append(Paragraph("Details", styles['Heading2']))
            details = data.get('details', [])
            
            for item in details:
                story.append(Paragraph(str(item), styles['Normal']))
            
            # Build PDF
            doc.build(story)
            return True
            
        except Exception as e:
            print(f"Error exporting to PDF: {e}")
            return False
    
    def export_report(
        self,
        data: Dict[str, Any],
        filename: str,
        format_type: str
    ) -> bool:
        """
        Export report to specified format.
        
        Args:
            data: Report data
            filename: Output filename
            format_type: "pdf", "excel", or "csv"
            
        Returns:
            True if successful
        """
        if format_type.lower() == "pdf":
            return self.export_to_pdf(data, filename)
        elif format_type.lower() in ["excel", "xlsx"]:
            return self.export_to_excel(data, filename)
        elif format_type.lower() == "csv":
            return self.export_to_csv(data, filename)
        else:
            print(f"Unknown format: {format_type}")
            return False
